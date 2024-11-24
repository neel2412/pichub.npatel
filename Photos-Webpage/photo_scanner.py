import os
import openpyxl
from PIL import Image
from PIL.ExifTags import TAGS
import numpy as np
from collections import Counter
from sklearn.cluster import KMeans

# Define paths
root_directory = "vintage/photos/"  # Update this path to your photo directory
excel_file = "photo_database.xlsx"

# Initialize Excel file
def initialize_excel(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Photos"
        ws.append(["Photo Path", "Keywords"])  # Header
        wb.save(file_path)

# Extract metadata from image (camera model, exposure time, etc.) using Pillow
def extract_metadata(image_path):
    try:
        # Open the image
        image = Image.open(image_path)

        # Extract EXIF data if available
        exif_data = image._getexif()
        if exif_data is not None:
            # Map EXIF tags to human-readable keys
            metadata_info = {}
            for tag, value in exif_data.items():
                tag_name = TAGS.get(tag, tag)
                metadata_info[tag_name] = value
        else:
            metadata_info = {}

        # Extract desired metadata from EXIF data (camera model, exposure time, orientation)
        camera_model = metadata_info.get("Model", "Unknown Camera")
        exposure_time = metadata_info.get("ExposureTime", "Unknown Exposure")
        orientation = metadata_info.get("Orientation", "Unknown Orientation")

        # Combine all metadata into a dictionary
        metadata_info = {
            "Camera": camera_model,
            "Exposure": exposure_time,
            "Orientation": orientation
        }
        return metadata_info
    except Exception as e:
        print(f"Error extracting metadata for {image_path}: {e}")
        return {}

# Extract dominant color from the image using KMeans clustering
def extract_dominant_color(image_path):
    try:
        # Open the image
        image = Image.open(image_path).convert("RGB")
        
        # Resize for faster processing
        image = image.resize((100, 100))

        # Convert image to numpy array and reshape for clustering
        img_data = np.array(image).reshape((-1, 3))

        # Perform KMeans clustering to find dominant colors
        kmeans = KMeans(n_clusters=1, random_state=0).fit(img_data)
        dominant_color = kmeans.cluster_centers_[0]
        
        # Convert RGB to a more human-readable format (e.g., hex)
        dominant_color_hex = '#{0:02x}{1:02x}{2:02x}'.format(int(dominant_color[0]), int(dominant_color[1]), int(dominant_color[2]))
        return dominant_color_hex
    except Exception as e:
        print(f"Error extracting color for {image_path}: {e}")
        return "Unknown Color"

# Generate keywords from image characteristics (metadata and color)
def extract_keywords(image_path):
    try:
        # Extract image metadata
        metadata_info = extract_metadata(image_path)
        
        # Extract dominant color
        dominant_color = extract_dominant_color(image_path)
        
        # Combine metadata and color into keywords
        keywords = [metadata_info.get("Camera", "Unknown Camera"),
                    metadata_info.get("Exposure", "Unknown Exposure"),
                    metadata_info.get("Orientation", "Unknown Orientation"),
                    "color: " + dominant_color]

        return keywords
    except Exception as e:
        print(f"Error extracting keywords for {image_path}: {e}")
        return ["error"]

# Scan directories and save metadata
def scan_and_save_photos(directory, excel_path, overwrite_existing=False):
    initialize_excel(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Photos"]

    # Read existing data from the Excel sheet
    existing_data = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2, values_only=True)}

    photo_count = 0  # Track the number of photos processed
    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):  # Check for image extensions
                file_path = os.path.join(root, file)
                
                # Check if photo is already indexed
                if file_path in existing_data and not overwrite_existing:
                    print(f"Skipping already indexed: {file_path}")
                    continue

                # Generate keywords for the photo
                keywords = extract_keywords(file_path)

                # Update or append to Excel
                if file_path in existing_data:
                    # Update existing entry if overwriting is allowed
                    if overwrite_existing:
                        print(f"Updating keywords for: {file_path}")
                        for row in ws.iter_rows(min_row=2):
                            if row[0].value == file_path:
                                row[1].value = ", ".join(keywords)
                else:
                    # Append new entry
                    ws.append([file_path, ", ".join(keywords)])
                    print(f"Added: {file_path}")
                
                photo_count += 1

    # Save updated Excel file
    wb.save(excel_path)
    print(f"Excel file updated. Total photos processed: {photo_count}")

# Run script
if __name__ == "__main__":
    # Set overwrite_existing=True to update keywords for existing photos
    scan_and_save_photos(root_directory, excel_file, overwrite_existing=False)
